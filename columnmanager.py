from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import openpyxl
import os


class ColumnManager:

    def __init__(self, file_location, sheet_name):
        self.workbook = self.load_workbook(file_location)
        self.sheet = self.get_sheet(self.workbook, sheet_name)
        self.file_location = file_location
        self.file_name = list(file_location.split("\\"))[-1]

    def load_workbook(self, file_location):
        """
        loads workbook object from excel file
        """
        try:
            return openpyxl.load_workbook(file_location)
        except:
            print("Problem laoding workbook from {}".format(file_location))

    def save_doc(self):  # Change to enable saving to any location *args
        """
        saves ColumnManager changes to excel workbook
        """
        self.workbook.save(self.file_location)

    def get_sheet(self, workbook_name, sheet_name):
        """
        retrieves sheet object from workbook object
        """
        try:
            return workbook_name[sheet_name]
        except:
            print("Problem finding {} worksheet from {} workbook".format(
                sheet_name, workbook_name))

    def make_new_column(self, title):
        """
        create column w/ given title
        """
        column_titles = self.get_column_titles()
        if title not in column_titles:
            try:
                max_column = self.sheet.max_column
                self.sheet.cell(row=1, column=max_column+1).value = title
            except:
                print("Something went wrong with building column {}".format(title))
        else:
            print("Column title {} already exists for file {}".format(title, self.file_name))

    def check_remove_title(self, list, exclude_title_toggle):  # maybe name this something else
        try:
            if exclude_title_toggle == True:
                return list[1:]
            elif exclude_title_toggle == False:
                return list
        except:
            print("Problem with exclude_title parameter. "
                  "Given value was {}".format(exclude_title))

    def get_column_titles(self, row=1):
        """
        retuns list with column title
        """
        try:
            max_column = self.sheet.max_column
            column_titles = [self.sheet.cell(row=row, column=column).value
                             for column in range(1, max_column+1)]
            return column_titles
        except:
            print("Problem retrieving column titles")

    def get_column_titles_with_index(self, row=1):
        """
        returns dict with column title and column index
        """
        max_column = self.sheet.max_column
        try:
            column_titles = {self.sheet.cell(row=row, column=column).value: column
                             for column in range(1, max_column+1)}
            return column_titles
        except:
            print("Problem retrieving column titles")

    def get_column_index(self, title, row=1):
        """
        return index number for given column title
        """
        try:
            columns = self.get_column_titles_with_index(row)
            return columns[title]
        except:
            print("Problem finding column titled {}".format(title))

    def get_column_cells(self, title, row=1):
        """
        returns list of cell objects from given column title
        """
        index = self.get_column_index(title, row)
        try:
            return [cell for cell in list(self.sheet.columns)[index - 1]]
        except:
            print("Problem finding cell for column w/ title {}".format(title))

    def get_column_cell_names(self, title, row=1):
        """
        return a list of cell names (ex."A32") for given column
        """
        cells = self.get_column_cells(title, row)
        try:
            cell_names = ["{}{}".format(cell.column, cell.row)
                          for cell in cells]
            return cell_names
        except:
            print("Problem building cell name from {} column".format(title))

    def get_column_values(self, title, row=1):
        """
        returns column values given title
        """
        try:
            cells = self.get_column_cells(title, row)
            values = [cell.value for cell in cells]
            return values
        except:
            print("Problem retrieving cells for column {}".format(title))

    def set_column_values(self, title, values, row_start=2):
        """
        sets cell values for given column
        """
        try:
            index = self.get_column_index(title)
            cells_to_set = self.get_column_cells(title)[1:]
        except:
            print("Problem with inputs to set_column_values")

        for i in range(len(cells_to_set)):
            try:
                self.sheet.cell(row=row_start, column=index).value = values[i]
            except:
                print("Problem setting cell values")
            row_start += 1

    def print_column_values(self, title, row=1):
        """
        prints out cell values
        """
        cells = self.get_column_cells(title, row)
        for cell in cells:
            print(cell.value)

    def divide_columns_values(self, first_title, second_title):  # maybe change return type to tuple
        """
        divides two columns and return list w/ resulting values
        """
        numerator = self.get_column_values(first_title)
        denominator = self.get_column_values(second_title)
        result = []
        for i in range(1, len(numerator)):
            try:
                result.append(numerator[i]/denominator[i])
            except (TypeError, ZeroDivisionError) as e:
                result.append("None")

        return result

    def divide_columns_formula(self, first_title, second_title):
        """
        return list of divided cell values from given columns
        """
        numerator_names = self.get_column_cell_names(first_title)
        denominator_names = self.get_column_cell_names(second_title)
        divided_formulas = ["={}/{}".format(numerator_names[i],
                                            denominator_names[i])
                            for i
                            in range(1, len(numerator_names))]
        return divided_formulas

    def gen_labordollar_perhour_column(self, with_formulas=True):
        # maybe this should be in another file
        """
        generates Labor $/Hour column with values
        """
        self.make_new_column("Labor $/Hr")
        if with_formulas == True:
            new_values = self.divide_columns_formula("Total Labor $",
                                                     "Total Hrs")
        elif with_formulas == False:
            new_values = self.divide_columns_values("Total Labor $", "Total Hrs")

        self.set_column_values("Labor $/Hr", new_values)

    def gen_laborhours_unitarea(self, with_formulas=True):
        # maybe this should be in another file
        """
        generates Labor Hours/Unit Area column w/ values
        """
        self.make_new_column("Labor Hours/Unit Area")
        if with_formulas == True:
            new_values = self.divide_columns_formula("Total Hrs", "Unit")
        elif with_formulas == False:
            new_values = self.divide_columns_values("Total Hrs", "Unit")

        self.set_column_values("Labor Hours/Unit Area", new_values)

    def color_column(self, title, color="B3FFB3"):
        """
        add background color to given column
        """
        cells = self.get_column_cells(title)
        for cell in cells:
            cell.fill = PatternFill(start_color=color,
                                    end_color=color,
                                    fill_type="solid")

    def add_validation(self, validation_string):
        self.sheet.insert_rows(1)
        dv = DataValidation(type="list", formula1=validation_string, allow_blank=True)
        self.sheet.add_data_validation(dv)
        cell = self.sheet["A1"]
        #cell.value = "Dog"
        dv.add(cell)

    def get_jobtype(self):
        return self.sheet.cell(row=1, column=1).value

    def show_error_location(self):
        print("Error found with file: {}".format(self.file_name.upper()))
